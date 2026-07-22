"""Persistencia real contra Dropbox (Fase 5). Implementa datasources.dropbox_base.

Un solo Excel (`registro.xlsx`) con cinco hojas append-only: NOVEDADES,
GASTOS_REALES, COBROS_DISTRIBUCION, TULAS_REPORTADAS, TARIFARIO.

Contrato de escritura:
  - APPEND-ONLY: descargar el workbook -> agregar UNA fila con openpyxl (jamas editar
    ni borrar filas existentes) -> subir con WriteMode.overwrite.
  - IDEMPOTENCIA: antes de agregar, releer la hoja y verificar la llave; si ya existe
    -> False (no se duplica).
  - detalle_json se guarda como string JSON en su celda.
  - Si el archivo no existe -> se crea con las 5 hojas + encabezados (y la carpeta
    /portal/ si hace falta) en la PRIMERA escritura.

TARIFARIO:
  - `leer_tarifario` es SOLO LECTURA: si la hoja tiene valores, sobreescriben los
    defaults del contrato (cambiar tarifas sin deploy); si el archivo/hoja no existe,
    devuelve los defaults (sin escribir). La hoja-plantilla editable con los defaults
    se materializa cuando el archivo se crea (primera escritura).

El cliente Dropbox (`dbx`) es INYECTABLE para poder testear con el SDK mockeado.
"""

from __future__ import annotations

import dataclasses
import io
import json
import posixpath

import config
from core.calculos import Tarifario
from datasources.dropbox_base import FuenteDropbox

# --------------------------------------------------------------------------- #
# Hojas y columnas (orden = orden de escritura de la fila)
# --------------------------------------------------------------------------- #

HOJA_NOVEDADES = "NOVEDADES"
HOJA_GASTOS = "GASTOS_REALES"
HOJA_COBROS = "COBROS_DISTRIBUCION"
HOJA_TULAS = "TULAS_REPORTADAS"
HOJA_TARIFARIO = "TARIFARIO"

COLS_NOVEDADES = ["fecha", "usuario", "manifiesto_id", "tula_codigo", "guia", "tipo",
                  "valor_esperado", "valor_real", "delta", "justificacion", "accion", "estado"]
COLS_GASTOS = ["fecha", "manifiesto_id", "proveedor", "concepto", "referencia",
               "detalle_json", "total_usd", "origen"]
COLS_COBROS = ["guia", "manifiesto_id", "transportadora", "lb_facturadas", "valor_usd"]
COLS_TULAS = ["manifiesto_id", "numero_tula", "kg_reportado", "awb_master", "fecha", "usuario", "origen"]
COLS_TARIFARIO = ["clave", "valor"]

_HOJAS = {
    HOJA_NOVEDADES: COLS_NOVEDADES,
    HOJA_GASTOS: COLS_GASTOS,
    HOJA_COBROS: COLS_COBROS,
    HOJA_TULAS: COLS_TULAS,
}


# --------------------------------------------------------------------------- #
# Llaves de idempotencia (funcionan sobre el objeto y sobre la fila releída)
# --------------------------------------------------------------------------- #

def _llave_novedad(d):
    return (d.get("manifiesto_id"), d.get("tula_codigo") or d.get("guia"), d.get("tipo"))


def _llave_gasto(d):
    return (d.get("manifiesto_id"), d.get("concepto"), d.get("referencia"))


def _llave_cobro(d):
    return (d.get("manifiesto_id"), d.get("guia"))


def _llave_tula(d):
    return (d.get("manifiesto_id"), d.get("numero_tula"))


def _es_not_found(exc) -> bool:
    try:
        return exc.error.is_path() and exc.error.get_path().is_not_found()
    except Exception:
        return "not_found" in str(exc).lower()


class FuenteDropboxAPI(FuenteDropbox):
    """Persistencia Dropbox real (ver docstring del módulo). `dbx` es inyectable."""

    def __init__(self, dbx, excel_path: str) -> None:
        self._dbx = dbx
        self._excel_path = excel_path

    # -- workbook: descarga / creación / subida ----------------------------- #

    def _descargar_workbook(self):
        import openpyxl
        from dropbox.exceptions import ApiError
        try:
            _, respuesta = self._dbx.files_download(self._excel_path)
        except ApiError as exc:
            if _es_not_found(exc):
                return None
            raise
        return openpyxl.load_workbook(io.BytesIO(respuesta.content))

    def _subir_workbook(self, wb) -> None:
        from dropbox.files import WriteMode
        buffer = io.BytesIO()
        wb.save(buffer)
        self._dbx.files_upload(buffer.getvalue(), self._excel_path, mode=WriteMode("overwrite"))

    def _asegurar_carpeta(self) -> None:
        from dropbox.exceptions import ApiError
        carpeta = posixpath.dirname(self._excel_path)
        if not carpeta or carpeta == "/":
            return
        try:
            self._dbx.files_create_folder_v2(carpeta)
        except ApiError:
            pass  # ya existe (conflict) — el upload igual funciona

    def _crear_workbook(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # quitar la hoja por defecto
        for hoja, columnas in _HOJAS.items():
            wb.create_sheet(hoja).append(columnas)
        # TARIFARIO como plantilla editable con el tarifario base (env/secret > DEMO).
        ws = wb.create_sheet(HOJA_TARIFARIO)
        ws.append(COLS_TARIFARIO)
        base = config.tarifario_base()
        for campo in dataclasses.fields(Tarifario):
            ws.append([campo.name, getattr(base, campo.name)])
        return wb

    def _asegurar_hojas(self, wb) -> None:
        """Para archivos legacy: crea cualquier hoja faltante con sus encabezados."""
        for hoja, columnas in _HOJAS.items():
            if hoja not in wb.sheetnames:
                wb.create_sheet(hoja).append(columnas)

    def _asegurar_workbook(self):
        """Workbook listo para escribir: descargado (con hojas completas) o nuevo."""
        wb = self._descargar_workbook()
        if wb is not None:
            self._asegurar_hojas(wb)
            return wb
        self._asegurar_carpeta()
        return self._crear_workbook()

    # -- escritura genérica (append-only + idempotencia) -------------------- #

    def _agregar(self, hoja, columnas, obj, llave_fn) -> bool:
        wb = self._asegurar_workbook()
        ws = wb[hoja]
        llave = llave_fn(obj)
        for fila in ws.iter_rows(min_row=2, values_only=True):
            existente = dict(zip(columnas, fila))
            if llave_fn(existente) == llave:
                return False  # ya existe -> no se duplica
        ws.append([_celda(obj.get(c)) for c in columnas])
        self._subir_workbook(wb)
        return True

    # -- lectura genérica --------------------------------------------------- #

    def _leer(self, hoja, columnas) -> list:
        wb = self._descargar_workbook()
        if wb is None or hoja not in wb.sheetnames:
            return []
        filas = []
        for fila in wb[hoja].iter_rows(min_row=2, values_only=True):
            if fila is None or all(v is None for v in fila):
                continue
            filas.append({c: v for c, v in zip(columnas, fila)})
        return filas

    # -- interfaz FuenteDropbox --------------------------------------------- #

    def leer_novedades(self) -> list:
        return self._leer(HOJA_NOVEDADES, COLS_NOVEDADES)

    def agregar_novedad(self, novedad) -> bool:
        return self._agregar(HOJA_NOVEDADES, COLS_NOVEDADES, novedad, _llave_novedad)

    def leer_gastos_reales(self) -> list:
        return self._leer(HOJA_GASTOS, COLS_GASTOS)

    def agregar_gasto_real(self, gasto) -> bool:
        return self._agregar(HOJA_GASTOS, COLS_GASTOS, gasto, _llave_gasto)

    def leer_cobros_distribucion(self) -> list:
        return self._leer(HOJA_COBROS, COLS_COBROS)

    def agregar_cobro_distribucion(self, cobro) -> bool:
        return self._agregar(HOJA_COBROS, COLS_COBROS, cobro, _llave_cobro)

    def leer_tulas_reportadas(self) -> list:
        return self._leer(HOJA_TULAS, COLS_TULAS)

    def agregar_tulas_reportadas(self, registro) -> bool:
        return self._agregar(HOJA_TULAS, COLS_TULAS, registro, _llave_tula)

    def leer_tarifario(self) -> Tarifario:
        """SOLO LECTURA. Precedencia: hoja TARIFARIO (Dropbox) > env/secret > default DEMO.
        La base (env/secret > DEMO) sale de config.tarifario_base(); la hoja, si existe,
        sobreescribe campo a campo."""
        base = config.tarifario_base()
        wb = self._descargar_workbook()
        if wb is None or HOJA_TARIFARIO not in wb.sheetnames:
            return base
        campos = {f.name for f in dataclasses.fields(Tarifario)}
        overrides = {}
        for fila in wb[HOJA_TARIFARIO].iter_rows(min_row=2, values_only=True):
            clave = fila[0] if fila else None
            valor = fila[1] if fila and len(fila) > 1 else None
            if clave is None or valor is None:
                continue
            clave = str(clave).strip()
            if clave in campos:
                try:
                    overrides[clave] = float(valor)
                except (TypeError, ValueError):
                    pass
        return dataclasses.replace(base, **overrides)


def _celda(valor):
    """Valor apto para celda de openpyxl. detalle_json ya viene como string JSON;
    listas/dicts residuales se serializan (defensa)."""
    if isinstance(valor, (list, dict)):
        return json.dumps(valor, ensure_ascii=False)
    return valor
