"""Tests de Fase 5: FuenteDropboxAPI con el SDK de Dropbox mockeado (sin red)."""

import io
import json

import openpyxl
import pytest

from datasources.dropbox_api import (COLS_NOVEDADES, FuenteDropboxAPI, HOJA_GASTOS,
                                     HOJA_NOVEDADES, HOJA_TARIFARIO)
from datasources.dropbox_base import FuenteDropbox

EXCEL_PATH = "/portal/registro.xlsx"

HOJAS_ESPERADAS = {"NOVEDADES", "GASTOS_REALES", "COBROS_DISTRIBUCION",
                   "TULAS_REPORTADAS", "TARIFARIO"}


def _nov(mid, guia, tipo="cobro_distribucion"):
    return {"fecha": "2026-07-16", "usuario": "admin1", "manifiesto_id": mid,
            "tula_codigo": None, "guia": guia, "tipo": tipo, "valor_esperado": 8,
            "valor_real": 12.5, "delta": 4.5, "justificacion": "x", "accion": "y",
            "estado": "resuelta"}


def _wb(dbx):
    return openpyxl.load_workbook(io.BytesIO(dbx.store[EXCEL_PATH]))


@pytest.fixture
def dbx(FakeDropbox):
    return FakeDropbox()


@pytest.fixture
def api(dbx):
    return FuenteDropboxAPI(dbx, EXCEL_PATH)


def test_cumple_la_interfaz(api):
    assert isinstance(api, FuenteDropbox)


# --- Creación del archivo con hojas + encabezados ------------------------- #

def test_crea_archivo_con_hojas_y_carpeta(api, dbx):
    assert dbx.store == {}                       # aún no existe
    assert api.agregar_novedad(_nov("M1", "G1")) is True

    wb = _wb(dbx)
    assert set(wb.sheetnames) == HOJAS_ESPERADAS
    assert [c.value for c in wb[HOJA_NOVEDADES][1]] == COLS_NOVEDADES   # encabezados
    assert wb[HOJA_NOVEDADES].max_row == 2                             # 1 fila de datos
    assert "/portal" in dbx.folders                                   # carpeta creada


# --- Append-only: no toca filas previas ----------------------------------- #

def test_append_no_toca_filas_previas(api, dbx):
    api.agregar_novedad(_nov("M1", "G1"))
    api.agregar_novedad(_nov("M1", "G2"))
    ws = _wb(dbx)[HOJA_NOVEDADES]
    idx = COLS_NOVEDADES.index("guia")
    guias = [fila[idx] for fila in ws.iter_rows(min_row=2, values_only=True)]
    assert guias == ["G1", "G2"]     # ambas, en orden, ninguna borrada
    assert ws.max_row == 3


# --- Idempotencia --------------------------------------------------------- #

def test_idempotencia_novedad(api):
    n = _nov("M1", "G1")
    assert api.agregar_novedad(n) is True
    assert api.agregar_novedad(dict(n)) is False           # misma llave
    assert api.agregar_novedad(_nov("M1", "G2")) is True   # otra guía


def test_idempotencia_gasto(api):
    g = {"manifiesto_id": "M1", "concepto": "importacion", "referencia": "F1",
         "detalle_json": "{}", "total_usd": 100.0, "proveedor": "X", "origen": "manual", "fecha": "2026-07-16"}
    assert api.agregar_gasto_real(g) is True
    assert api.agregar_gasto_real(dict(g)) is False
    assert api.agregar_gasto_real({**g, "referencia": "F2"}) is True


def test_idempotencia_cobro_y_tula(api):
    c = {"manifiesto_id": "M1", "guia": "G1", "transportadora": "T", "lb_facturadas": 5, "valor_usd": 12.5}
    assert api.agregar_cobro_distribucion(c) is True
    assert api.agregar_cobro_distribucion(dict(c)) is False
    t = {"manifiesto_id": "M1", "numero_tula": 3, "kg_reportado": 20.0, "awb_master": "999-1",
         "fecha": "2026-07-16", "usuario": "j", "origen": "vision"}
    assert api.agregar_tulas_reportadas(t) is True
    assert api.agregar_tulas_reportadas(dict(t)) is False


def test_tula_reportada_persiste_awb_master(api):
    t = {"manifiesto_id": "900009", "numero_tula": 1, "kg_reportado": 38.6,
         "awb_master": "999-11112222", "fecha": "2026-07-17", "usuario": "j", "origen": "vision"}
    assert api.agregar_tulas_reportadas(t) is True
    leidos = api.leer_tulas_reportadas()
    assert leidos[0]["awb_master"] == "999-11112222"        # el vínculo manifiesto↔awb persiste
    assert leidos[0]["numero_tula"] == 1


# --- detalle_json como string en la celda + lectura ----------------------- #

def test_detalle_json_string_y_lectura(api):
    detalle = json.dumps({"kg_cobrados": 774, "tarifa_kg_usd": 1.00})
    gasto = {"fecha": "2026-07-16", "manifiesto_id": "900007", "proveedor": "DemoAir",
             "concepto": "importacion", "referencia": "DEMO-0001", "detalle_json": detalle,
             "total_usd": 834.00, "origen": "vision"}
    assert api.agregar_gasto_real(gasto) is True
    leidos = api.leer_gastos_reales()
    assert len(leidos) == 1
    assert leidos[0]["detalle_json"] == detalle                       # string, no dict
    assert json.loads(leidos[0]["detalle_json"])["kg_cobrados"] == 774


# --- TARIFARIO: sin hoja (defaults, read-only) y con hoja (override) ------- #

def test_tarifario_sin_archivo_defaults_sin_escribir(api, dbx):
    t = api.leer_tarifario()
    assert t.imp_tarifa_kg == 1.00 and t.imp_awb_fee == 10.0
    assert dbx.store == {}                       # leer_tarifario NO escribe


def test_tarifario_override_desde_hoja(api, dbx):
    api.agregar_novedad(_nov("M1", "G1"))        # crea el archivo con TARIFARIO plantilla
    wb = _wb(dbx)
    for fila in wb[HOJA_TARIFARIO].iter_rows(min_row=2):
        if fila[0].value == "imp_tarifa_kg":
            fila[1].value = 1.40
    buf = io.BytesIO(); wb.save(buf); dbx.store[EXCEL_PATH] = buf.getvalue()

    t = api.leer_tarifario()
    assert t.imp_tarifa_kg == 1.40               # override aplicado
    assert t.imp_awb_fee == 10.0                 # el resto sigue default


def test_lecturas_vacias_si_no_existe(api):
    assert api.leer_novedades() == []
    assert api.leer_gastos_reales() == []
    assert api.leer_cobros_distribucion() == []
    assert api.leer_tulas_reportadas() == []
